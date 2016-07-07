#   Copyright 2016 NICTA
#
#   Licensed under the Apache License, Version 2.0 (the "License");
#   you may not use this file except in compliance with the License.
#   You may obtain a copy of the License at
#
#       http://www.apache.org/licenses/LICENSE-2.0
#
#   Unless required by applicable law or agreed to in writing, software
#   distributed under the License is distributed on an "AS IS" BASIS,
#   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#   See the License for the specific language governing permissions and
#   limitations under the License.


import datetime
import csv
from scipy import stats
from decimal import Decimal, ROUND_HALF_UP
import re
from openpyxl import load_workbook
from dashboard_loader.loader_utils import *
from widget_def.models import Parametisation, ParametisationValue
from coag_uploader.models import *

# These are the names of the groups that have permission to upload data for this uploader.
# If the groups do not exist they are created on registration.
groups = [ "upload_all", "upload_coag" ]

# This describes the file format.  It is used to describe the format by both
# "python manage.py upload_data frontlineservice_uploader" and by the uploader 
# page in the data admin GUI.
file_format = {
    "format": "xlsx",
    "sheets": [
         {
             "name": "Table of Contents",
             "rows": [],
             "cols": [],
             "notes": [ "Not read by uploader", ],
         },
         {
             "name": "1 NAHA",
             "rows": [],
             "cols": [],
             "notes": [ "TODO", ],
         },
         {
             "name": "2 NAHA",
             "rows": [],
             "cols": [],
             "notes": [ "TODO", ],
         },
         {
             "name": "3 NAHA",
             "rows": [],
             "cols": [],
             "notes": [ "TODO", ],
         },
         {
             "name": "4 NAHA",
             "rows": [],
             "cols": [],
             "notes": [ "TODO", ],
         },
    ],
}

def upload_file(uploader, fh, actual_freq_display=None, verbosity=0):
    messages = []
    try:
        if verbosity > 0:
            messages.append("Loading workbook...")
        wb = load_workbook(fh, read_only=True)
        messages.extend(load_housing_data(wb, verbosity))
    except LoaderException, e:
        raise e
    except Exception, e:
        raise LoaderException("Invalid file: %s" % unicode(e))
    return messages

def load_housing_data(wb, verbosity):
    messages = []
    if verbosity > 1:
        messages.append("Loading Housing Data...")
    messages.extend(load_housing_rental_stress(wb, verbosity))
    messages.extend(load_housing_homelessness(wb, verbosity))
    messages.extend(load_housing_indigenous_ownership(wb, verbosity))
    messages.extend(load_housing_indigenous_crowding(wb, verbosity))
    return messages

def column_labels(mini, maxi):
    labels=[]
    alphabet = [ chr(i) for i in range(ord('A'), ord('A') + 26) ]
    alphabetp = [''] + alphabet
    ri = 0
    for c0 in alphabetp:
        for c1 in alphabet:
            ri += 1
            col = c0 + c1
            if ri >= mini and ri <= maxi:
                labels.append(col)
            if ri > maxi:
                return labels

def find_state_cols(sheet, row):
    state_cols = {}
    for col in column_labels(sheet.min_column, sheet.max_column):
       val = sheet["%s%d" % (col, row)].value
       if val in state_map:
            state_cols[state_map[val]] = col
    return state_cols

def zero_all_rows(rows):
    for k in rows.keys():
        rows[k] = 0

def all_rows_found(rows):
    for k in rows.keys():
        if not rows[k]:
            return False
    return True

def load_state_grid(wb, sheet_name, data_category, dataset, abort_on, model, first_cell_rows, intermediate_cell_rows, verbosity):
    messages = []
    if verbosity > 2:
        messages.append("Loading %s Data: %s" % (data_category, dataset))
    sheet = wb[sheet_name]
    state_cols = {}
    row = 1
    while not state_cols:
        state_cols = find_state_cols(sheet, row)
        row += 1
        if not state_cols and row > sheet.max_row:
            raise LoaderException("No State Columns found in worksheet '%s'" % sheet_name)
    year = 0
    isfy = False
    rows = {}
    for fld in first_cell_rows.keys():
        rows[fld] = 0
    for fld in intermediate_cell_rows.keys():
        rows[fld] = 0
    while True:
        first_cell=sheet["A%d" % row].value
        if first_cell:
            if first_cell == abort_on:
                break
            first_cell_matched = False
            for fld, fcr in first_cell_rows.items():
                if first_cell == fcr:
                    first_cell_matched=True
                    rows[fld] = row
                    break
            if not first_cell_matched:
                try:
                    (year, isfy) = parse_year(first_cell)
                    zero_all_rows(rows)
                except:
                    pass
        if year:
            for col in column_labels(2, sheet.max_column):
                if col in state_cols.values():
                    break
                cval = sheet["%s%d" % (col, row)].value
                for fld, icr in intermediate_cell_rows.items():
                    if cval == icr:
                        rows[fld]=row
                        break
                if year and all_rows_found(rows):
                    for state, scol in state_cols.items():
                        defaults = { "financial_year": isfy }
                        for fld, frow in rows.items():
                            defaults[fld] = sheet["%s%d" % (scol, frow)].value
                        obj, created = model.objects.update_or_create(
                                    state=state,
                                    year=year,
                                    defaults=defaults)
                    zero_all_rows(rows)
        row += 1
    return messages

def flt_year_equal(y1,y2):
    return abs(y1-y2) < 1e-6

def current_float_year():
    today=datetime.date.today()
    flt_year = float(today.year)
    if today.month >= 7:
        flt_year += 0.5
    return flt_year

def find_trend(metrics, reference_year, target_year):
    reference = None
    datum = None
    years = []
    data = []
    for metric in metrics:
        years.append(metric[0])
        data.append(metric[1])
        if flt_year_equal(metric[0], reference_year):
            reference = metric[1]
        if flt_year_equal(metric[0], target_year):
            datum = metric[1]
    if reference is None:
        return None
    if datum is None:
        slope, intercept, r_val, p_val, std_err = stats.linregress(years,data)
        datum = slope * target_year + intercept
        projected = True
    else:
        projected = False
    return {
            "datum": datum, 
            "reference": reference, 
            "benchmark": (datum-reference)/reference,
            "projected": projected
    }

def calculate_benchmark(reference_year, target_year, 
                    target, desire_overtarget, 
                    model, benchmark_field, 
                    widget_url, widget_label, stat_base_lbl,
                    verbosity):
    messages = []
    metrics = [ (obj.float_year(), float(getattr(obj,benchmark_field))) for obj in model.objects.filter(state=AUS).order_by("year", "financial_year") ]
    trend = find_trend(metrics, reference_year, target_year)
    if trend:
        benchmark = trend["benchmark"]*100.0
        datum = trend["datum"]
        reference = trend["reference"]
        if (desire_overtarget and trend["benchmark"] >= target) or (not desire_overtarget and trend["benchmark"] <= target):
            if trend["projected"]:
                if target_year >= current_float_year():
                    tlc = "likely_to_have_been_met"
                else:
                    tlc = "on_track"
            else:
                tlc = "achieved"
        else:
            if trend["projected"]:
                if target_year >= current_float_year():
                    tlc = "unlikely_to_have_been_met"
                else:
                    tlc = "not_on_track"
            else:
                tlc = "not_met"
        if trend["projected"]:
            outcome_type = "projection"     
        else:
            outcome_type = "complete"     
    else:
        benchmark = 0.0
        reference = 0.0
        datum = 0.0
        outcome_type = "N/A"
        tlc = "new_revised_benchmark"
    set_statistic_data(widget_url, widget_label, "outcome", 
                    abs(benchmark), traffic_light_code=tlc, 
                    trend=cmp(datum, reference))
    set_statistic_data(widget_url, widget_label, "outcome_type", 
                    outcome_type)
    set_statistic_data(widget_url, widget_label, "benchmark", 
                    abs(target*100.0), trend=cmp(target, 0.0))
    set_statistic_data(widget_url, widget_label, "benchmark_year", 
                    display_float_year(target_year))
    clear_statistic_list(widget_url, widget_label, "data")
    add_statistic_list_item(widget_url, widget_label, "data",
                            reference, 10,
                            label=display_float_year(reference_year))
    last_metric = metrics[-1]
    add_statistic_list_item(widget_url, widget_label, "data",
                            last_metric[1], 20,
                            label=display_float_year(last_metric[0]))
    return messages

def load_housing_rental_stress(wb, verbosity):
    messages = []
    messages.extend(load_state_grid(wb, "1 NAHA", 
                        "Housing", "Rental Stress", 
                        "Notes:", HousingRentalStressData,
                        {}, {"percentage": "%", "uncertainty": "+"}, 
                        verbosity))
    messages.extend(calculate_benchmark(2007.5, 2015.5, 
                            -0.1, False, 
                            HousingRentalStressData, "percentage", 
                            "rental_stress", "rental_stress", "percent",
                            verbosity))
    return messages


def load_housing_homelessness(wb, verbosity):
    messages = []
    messages.extend(load_state_grid(wb, "2 NAHA", 
                        "Housing", "Homelessness", 
                        "Notes:", HousingHomelessData,
                        {
                            "homeless_persons": "All homeless persons", 
                            "rate_per_10k": "Rate per 10,000 of the population"
                        }, 
                        { "percent_of_national": "%",}, verbosity))
    messages.extend(calculate_benchmark(2006.0, 2013.0,
                            -0.07, False,
                            HousingHomelessData, "homeless_persons",
                            "homelessness", "homelessness", "count",
                            verbosity))
    return messages

def load_housing_indigenous_ownership(wb, verbosity):
    messages = []
    messages.extend(load_state_grid(wb, "3 NAHA", 
                        "Housing", "Indigenous Ownership", 
                        "Notes:", IndigenousHomeOwnershipData,
                        { "uncertainty": "95 per cent confidence interval"}, 
                        { "percentage": "%"}, verbosity))
    messages.extend(calculate_benchmark(2008.0, 2017.5,
                            0.1, True,
                            IndigenousHomeOwnershipData, "percentage",
                            "indigenous_home_ownership", "indigenous_home_ownership", 
                            "percent", verbosity))
    return messages

def load_housing_indigenous_crowding(wb, verbosity):
    messages = []
    messages.extend(load_state_grid(wb, "4 NAHA", 
                        "Housing", "Indigenous Overcrowding", 
                        "Notes:", IndigenousOvercrowdingData,
                        {}, {"percentage": "%", "uncertainty": "+"}, 
                        verbosity))
    messages.extend(calculate_benchmark(2008.0, 2017.5,
                            -0.2, False,
                            IndigenousOvercrowdingData, "percentage",
                            "indigenous_overcrowding", "indigenous_overcrowding", 
                            "percent", verbosity))

    return messages

